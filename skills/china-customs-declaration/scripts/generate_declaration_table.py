#!/usr/bin/env python3
"""
报关资料表格生成脚本（增强版）

从商品数据生成：
  - 商品资料汇总表（含确认状态）
  - 规范申报要素表（基于Schema）
  - 商业发票草稿
  - 装箱单草稿
  - 待确认问题清单
  - 风险问题清单（基于条件式风险引擎）

支持格式：XLSX, CSV, JSON
"""

import csv
import json
import os
import sys
from datetime import datetime
from decimal import Decimal
from typing import Any

# Add common modules to path
_script_dir = os.path.dirname(os.path.abspath(__file__))
if _script_dir not in sys.path:
    sys.path.insert(0, _script_dir)

from common.decimal_utils import (
    format_amount,
    round_amount,
    to_decimal,
    to_decimal_safe,
)
from common.field_normalizer import (
    FIELD_ALIASES,
    normalize_hs_code,
)

# Optional: declaration_elements module
try:
    import declaration_elements as de

    HAS_DECL_MODULE = True
except ImportError:
    HAS_DECL_MODULE = False

# Optional: regulatory_risk_engine module
try:
    import regulatory_risk_engine as rre

    HAS_RISK_MODULE = True
except ImportError:
    HAS_RISK_MODULE = False

# ──────────────────────────────────────────────
# Data Loading
# ──────────────────────────────────────────────


def load_data_from_xlsx(filepath: str) -> list[dict]:
    """从 XLSX 文件加载数据，支持多 Sheet 自动检测和字段别名解析"""
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("需要安装 openpyxl: pip install openpyxl")
        return []

    wb = load_workbook(filepath, data_only=True)

    # 构建全量已知字段名（大小写/空格/下划线归一化）
    known_fields: set[str] = set()
    for aliases in FIELD_ALIASES.values():
        for a in aliases:
            known_fields.add(a.lower().replace(" ", "").replace("_", ""))

    best_sheet = None
    best_score = 0
    best_headers: list[str] = []
    best_header_row = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        # 尝试前 3 行作为潜在表头
        for header_row_idx in range(min(3, len(rows))):
            headers = [
                str(c).strip() if c is not None else "" for c in rows[header_row_idx]
            ]
            score = 0
            for h in headers:
                h_clean = h.lower().replace(" ", "").replace("_", "")
                if h_clean in known_fields:
                    score += 1
            if score > best_score:
                best_score = score
                best_sheet = sheet_name
                best_headers = headers
                best_header_row = header_row_idx

    if best_score == 0:
        print(
            f"错误: 无法在文件中识别数据表头。支持的字段包括: {list(FIELD_ALIASES.keys())}"
        )
        return []

    ws = wb[best_sheet]
    rows = list(ws.iter_rows(values_only=True))
    data_rows = rows[best_header_row + 1 :]

    # 检查合并单元格并警告
    try:
        merged_cells = list(ws.merged_cells.ranges)
        if merged_cells:
            print(
                f"提示: 工作表 '{best_sheet}' 包含 {len(merged_cells)} 个合并单元格（已取左上角值）"
            )
    except Exception:
        pass

    result: list[dict] = []
    empty_count = 0
    for row_idx, row in enumerate(data_rows):
        values = [str(c).strip() if c is not None else "" for c in (row if row else [])]
        if all(v == "" for v in values):
            empty_count += 1
            continue
        record: dict[str, str] = {}
        for i, header in enumerate(best_headers):
            if header and i < len(values):
                record[header] = values[i]
        if any(v.strip() for v in record.values()):
            result.append(record)

    if empty_count > 0:
        print(f"提示: 跳过了 {empty_count} 个空行")

    return result


def load_data(filepath: str) -> list[dict]:
    """加载数据，支持 JSON / CSV / XLSX 格式"""
    ext = os.path.splitext(filepath)[1].lower()

    if ext == ".json":
        with open(filepath, "r", encoding="utf-8") as f:
            data = json.load(f)
        if isinstance(data, list):
            return data
        if isinstance(data, dict):
            return data.get("products", data.get("items", data.get("data", [data])))
        return []

    elif ext == ".csv":
        data: list[dict] = []
        with open(filepath, "r", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            for row in reader:
                cleaned = {k.strip(): v.strip() if v else "" for k, v in row.items()}
                data.append(cleaned)
        return data

    elif ext in (".xlsx", ".xls"):
        return load_data_from_xlsx(filepath)

    else:
        print(f"错误: 不支持的文件格式 '{ext}'。仅支持 JSON, CSV, XLSX 格式。")
        return []


# ──────────────────────────────────────────────
# Table Generators
# ──────────────────────────────────────────────


def _resolve_verification_status(row: dict) -> str:
    """根据数据完整度判断确认状态"""
    required_fields = ["HS编码", "品牌", "型号", "用途"]
    missing = [f for f in required_fields if not row.get(f, "").strip()]
    if not missing:
        return "已确认"
    if len(missing) <= 1:
        return "基本确认"
    return "待确认"


def _row_display_value(row: dict, key: str) -> str:
    """获取行字段的显示值，统一返回字符串"""
    v = row.get(key, "")
    if v is None:
        return ""
    return str(v).strip()


def _row_decimal(row: dict, key: str, default: str = "") -> str:
    """获取行字段的 Decimal 格式化值"""
    v = to_decimal(row.get(key))
    if v is None:
        return default
    return str(format_amount(v))


def generate_summary_table(data: list[dict]) -> list[dict]:
    """生成商品资料汇总表（含确认状态）"""
    return [
        {
            "项号": i,
            "中文品名": _row_display_value(row, "中文品名"),
            "英文品名": _row_display_value(row, "英文品名"),
            "品牌": _row_display_value(row, "品牌"),
            "型号": _row_display_value(row, "型号"),
            "HS编码": normalize_hs_code(row.get("HS编码", "")),
            "用途": _row_display_value(row, "用途"),
            "功能": _row_display_value(row, "功能"),
            "工作原理": _row_display_value(row, "工作原理"),
            "材质": _row_display_value(row, "材质"),
            "技术参数": _row_display_value(row, "技术参数"),
            "是否整机": _row_display_value(row, "是否整机"),
            "是否含无线": _row_display_value(row, "是否含无线"),
            "是否含电池": _row_display_value(row, "是否含电池"),
            "数量": _row_display_value(row, "数量"),
            "单位": _row_display_value(row, "单位"),
            "单价": _row_decimal(row, "单价"),
            "总价": _row_decimal(row, "总价"),
            "币种": _row_display_value(row, "币种"),
            "原产国": _row_display_value(row, "原产国"),
            "确认状态": _resolve_verification_status(row),
        }
        for i, row in enumerate(data, 1)
    ]


def generate_invoice(data: list[dict]) -> tuple[list[dict], dict]:
    """生成商业发票草稿（使用 Decimal 计算）"""
    today = datetime.now().strftime("%Y-%m-%d")
    total_amount = Decimal("0")

    invoice_rows = []
    for i, row in enumerate(data, 1):
        qty = to_decimal_safe(row.get("数量"))
        unit_price = to_decimal_safe(row.get("单价"))
        total_price = round_amount(qty * unit_price)
        total_amount += total_price

        invoice_rows.append(
            {
                "项号": i,
                "中文品名": _row_display_value(row, "中文品名"),
                "英文品名": _row_display_value(row, "英文品名"),
                "品牌型号": f"{_row_display_value(row, '品牌')} {_row_display_value(row, '型号')}".strip(),
                "HS编码": normalize_hs_code(row.get("HS编码", "")),
                "数量": _row_display_value(row, "数量"),
                "单位": _row_display_value(row, "单位"),
                "单价": format_amount(unit_price),
                "总价": format_amount(total_price),
                "币种": _row_display_value(row, "币种") or "USD",
                "原产国": _row_display_value(row, "原产国"),
            }
        )

    meta = {
        "发票日期": today,
        "总金额": format_amount(total_amount),
    }
    return invoice_rows, meta


def generate_packing_list(data: list[dict]) -> list[dict]:
    """生成装箱单草稿（使用 Decimal）"""
    total_net_weight = Decimal("0")
    total_gross_weight = Decimal("0")

    rows = []
    for i, row in enumerate(data, 1):
        net_w = to_decimal_safe(row.get("净重"))
        gross_w = to_decimal_safe(row.get("毛重"))
        total_net_weight += net_w
        total_gross_weight += gross_w

        rows.append(
            {
                "箱号": i,
                "品名": _row_display_value(row, "中文品名"),
                "型号": _row_display_value(row, "型号"),
                "数量": _row_display_value(row, "数量"),
                "包装单位": _row_display_value(row, "包装单位")
                or _row_display_value(row, "单位"),
                "净重(kg)": format_amount(net_w),
                "毛重(kg)": format_amount(gross_w),
                "包装尺寸(cm)": _row_display_value(row, "包装尺寸"),
                "体积(m³)": _row_display_value(row, "体积"),
            }
        )

    # 添加汇总行
    rows.append(
        {
            "箱号": "",
            "品名": "合计",
            "型号": "",
            "数量": "",
            "包装单位": "",
            "净重(kg)": format_amount(total_net_weight),
            "毛重(kg)": format_amount(total_gross_weight),
            "包装尺寸(cm)": "",
            "体积(m³)": "",
        }
    )

    return rows


# ──────────────────────────────────────────────
# Declaration Elements
# ──────────────────────────────────────────────


def generate_declaration_elements(data: list[dict]) -> tuple[list[dict], str]:
    """生成申报要素明细表

    Returns:
        (elements_table, source_info) — source_info 描述数据来源
    """
    if HAS_DECL_MODULE:
        try:
            registry = de.load_schema_registry()
            batch_results = de.generate_batch_elements(data, registry)
            return _format_schema_elements(
                data, batch_results
            ), "基于 Schema 匹配 (declaration_elements)"
        except Exception as e:
            print(
                f"警告: declaration_elements 模块执行异常 ({e})，回退到基础模式",
                file=sys.stderr,
            )

    # Fallback: 基础固定字段模式
    return _legacy_elements(data), "基础字段（回退模式，declaration_elements 不可用）"


def _format_schema_elements(data: list[dict], batch_results: list[dict]) -> list[dict]:
    """将 Schema 匹配结果格式化为表格行"""
    elements = []
    # 建立 HS 编码到结果的映射
    result_map: dict[str, dict] = {}
    for r in batch_results:
        hs = r.get("hs_code", "")
        if hs:
            result_map[hs] = r

    for i, row in enumerate(data, 1):
        hs_code = normalize_hs_code(row.get("HS编码", ""))
        decl_result = result_map.get(hs_code, {})

        schema_type = decl_result.get("schema_type", "not_found")
        verification_status = decl_result.get("verification_status", "待官方核实")
        fields = decl_result.get("fields", [])
        detail_version = decl_result.get("detail_version", [])

        if fields:
            for f in fields:
                elements.append(
                    {
                        "项号": i,
                        "HS编码": hs_code,
                        "品名": _row_display_value(row, "中文品名"),
                        "申报字段": f["name"],
                        "申报内容": f["content"],
                        "信息来源": f.get("source", ""),
                        "状态": f.get("status", verification_status),
                        "Schema类型": schema_type,
                    }
                )
        elif detail_version:
            for d in detail_version:
                elements.append(
                    {
                        "项号": i,
                        "HS编码": hs_code,
                        "品名": _row_display_value(row, "中文品名"),
                        "申报字段": d["申报要素"],
                        "申报内容": d["填报内容"],
                        "信息来源": "",
                        "状态": d.get("状态", verification_status),
                        "Schema类型": schema_type,
                    }
                )
        else:
            # Schema 存在但无字段定义
            elements.append(
                {
                    "项号": i,
                    "HS编码": hs_code,
                    "品名": _row_display_value(row, "中文品名"),
                    "申报字段": f"Schema ({schema_type})",
                    "申报内容": decl_result.get("paste_version", ""),
                    "信息来源": decl_result.get("source", ""),
                    "状态": verification_status,
                    "Schema类型": schema_type,
                }
            )
    return elements


def _legacy_elements(data: list[dict]) -> list[dict]:
    """基础固定字段申报要素（回退模式）"""
    elements = []
    for i, row in enumerate(data, 1):
        # 固定字段列表
        legacy_fields = [
            (
                "品牌类型",
                row.get("品牌类型", "待确认"),
                row.get("信息来源", ""),
                "待确认",
            ),
            ("出口享惠情况", row.get("出口享惠情况", "待确认"), "", "待确认"),
            (
                "用途",
                row.get("用途", "待确认"),
                row.get("信息来源", ""),
                "已确认" if row.get("用途") else "待确认",
            ),
            (
                "品牌",
                row.get("品牌", "待确认"),
                row.get("信息来源", ""),
                "已确认" if row.get("品牌") else "待确认",
            ),
            (
                "型号",
                row.get("型号", "待确认"),
                row.get("信息来源", ""),
                "已确认" if row.get("型号") else "待确认",
            ),
            ("其他", row.get("其他", "待确认"), "", "待确认"),
        ]
        for field_name, content, source, status in legacy_fields:
            elements.append(
                {
                    "项号": i,
                    "HS编码": normalize_hs_code(row.get("HS编码", "")),
                    "品名": _row_display_value(row, "中文品名"),
                    "申报字段": field_name,
                    "申报内容": content,
                    "信息来源": source,
                    "状态": status,
                }
            )
    return elements


# ──────────────────────────────────────────────
# Pending Questions
# ──────────────────────────────────────────────


def generate_pending_questions(data: list[dict]) -> list[str]:
    """生成待确认问题清单"""
    questions = []
    for i, row in enumerate(data, 1):
        prefix = f"第{i}项 ({_row_display_value(row, '中文品名')}): "
        if not row.get("用途"):
            questions.append(f"{prefix}产品的主要用途是什么（家用/商用/工业等）？")
        if not row.get("功能"):
            questions.append(f"{prefix}产品的主要功能是什么？")
        if not row.get("工作原理"):
            questions.append(f"{prefix}产品的工作原理是什么？")
        if not row.get("是否含无线"):
            questions.append(f"{prefix}产品是否含WiFi、蓝牙等无线功能？")
        if not row.get("是否含电池"):
            questions.append(f"{prefix}产品是否含锂电池？如有，请提供容量（mAh/Wh）")
        if not row.get("是否整机"):
            questions.append(f"{prefix}产品是整机还是零件/附件？")
        if not row.get("原产国"):
            questions.append(f"{prefix}产品的原产国是哪里？")
    return questions


# ──────────────────────────────────────────────
# Risk Detection
# ──────────────────────────────────────────────


def detect_risks(data: list[dict]) -> tuple[list[dict], str]:
    """检测风险项目

    Returns:
        (risks, engine_info) — engine_info 描述风险引擎来源
    """
    if HAS_RISK_MODULE:
        try:
            all_risks = []
            for i, row in enumerate(data, 1):
                # 将产品行转换为 risk engine 期望的格式
                product_data = _row_to_rre_format(row)
                risk_results = rre.evaluate_risks(product_data)
                for r in risk_results:
                    r["问题位置"] = f"第{i}项 ({_row_display_value(row, '中文品名')})"
                all_risks.extend(risk_results)

            if all_risks:
                return (
                    all_risks,
                    f"基于规则引擎 (regulatory_risk_engine) — {len(all_risks)} 项风险",
                )
            return all_risks, "基于规则引擎 (regulatory_risk_engine) — 未发现风险"
        except Exception as e:
            print(
                f"警告: regulatory_risk_engine 执行异常 ({e})，回退到基础检测",
                file=sys.stderr,
            )

    # Fallback: 基础检测
    return _legacy_risks(data), "基础检测（回退模式）"


def _row_to_rre_format(row: dict) -> dict:
    """将表格行数据转换为 risk engine 期望的 dict 格式"""
    return {
        "name_cn": _row_display_value(row, "中文品名"),
        "name_en": _row_display_value(row, "英文品名"),
        "brand": _row_display_value(row, "品牌"),
        "model": _row_display_value(row, "型号"),
        "hs_code": normalize_hs_code(row.get("HS编码", "")),
        "usage": _row_display_value(row, "用途"),
        "function": _row_display_value(row, "功能"),
        "material": _row_display_value(row, "材质"),
        "has_wireless": _row_display_value(row, "是否含无线").lower()
        in ("是", "yes", "true", "含"),
        "has_battery": _row_display_value(row, "是否含电池").lower()
        in ("是", "yes", "true", "含"),
    }


def _legacy_risks(data: list[dict]) -> list[dict]:
    """基础风险检测（回退模式）"""
    risks = []
    for i, row in enumerate(data, 1):
        prefix = f"第{i}项 ({_row_display_value(row, '中文品名')})"

        # 含无线功能但未提示 SRRC
        if _row_display_value(row, "是否含无线").lower() in ("是", "yes", "true", "含"):
            risks.append(
                {
                    "风险等级": "verify",
                    "问题位置": prefix,
                    "问题说明": "含无线功能，需确认是否取得SRRC型号核准证",
                    "修改建议": "确认SRRC认证状态，如未取得需办理",
                }
            )

        # 含电池但未提示
        if _row_display_value(row, "是否含电池").lower() in ("是", "yes", "true", "含"):
            risks.append(
                {
                    "风险等级": "verify",
                    "问题位置": prefix,
                    "问题说明": "含电池，需确认UN38.3/MSDS/危包证及CCC认证状态",
                    "修改建议": "准备UN38.3报告、MSDS、危包证，确认CCC证书",
                }
            )

        # 品名过宽
        pname = _row_display_value(row, "中文品名")
        if pname in ("电子产品", "设备", "配件", "零件", "机器", "仪器"):
            risks.append(
                {
                    "风险等级": "medium",
                    "问题位置": prefix,
                    "问题说明": f"品名'{pname}'过于宽泛，建议具体化",
                    "修改建议": "改为具体品名，如'激光投影仪'、'蓝牙音箱'等",
                }
            )

        # 缺少型号
        if not _row_display_value(row, "型号"):
            risks.append(
                {
                    "风险等级": "medium",
                    "问题位置": prefix,
                    "问题说明": "型号缺失，可能影响申报",
                    "修改建议": "补充产品型号",
                }
            )

    return risks


# ──────────────────────────────────────────────
# Output Functions
# ──────────────────────────────────────────────


def save_as_xlsx(tables: dict[str, Any], output_path: str) -> None:
    """保存为 Excel 文件（多 Sheet）"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        print("需要安装 openpyxl: pip install openpyxl")
        return

    wb = Workbook()
    wb.remove(wb.active)

    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(
        start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"
    )

    for sheet_name, table_data in tables.items():
        if not table_data:
            continue
        ws = wb.create_sheet(title=sheet_name[:31])  # Excel sheet name limit

        if isinstance(table_data, list):
            if not table_data:
                continue
            headers = list(table_data[0].keys())
            # 写表头
            ws.append(headers)
            for col_idx, _ in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")

            # 写数据
            for row in table_data:
                ws.append([row.get(h, "") for h in headers])

            # 自动列宽
            for col_idx, header in enumerate(headers, 1):
                col_width = max(len(str(header)) * 2, 10)
                ws.column_dimensions[
                    chr(64 + col_idx) if col_idx <= 26 else "A"
                ].width = min(col_width, 40)

        elif isinstance(table_data, dict):
            for key, value in table_data.items():
                ws.append([key, value])

    wb.save(output_path)
    print(f"Excel 已保存到: {output_path}")


def save_as_csv(data: list[dict], output_path: str) -> None:
    """保存为 CSV"""
    if not data:
        print(f"数据为空，跳过: {output_path}")
        return
    with open(output_path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=data[0].keys())
        writer.writeheader()
        writer.writerows(data)
    print(f"CSV 已保存到: {output_path}")


def save_as_json(data: Any, output_path: str) -> None:
    """保存为 JSON"""
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f"JSON 已保存到: {output_path}")


# ──────────────────────────────────────────────
# Sample Data
# ──────────────────────────────────────────────

SAMPLE_DATA = [
    {
        "中文品名": "家用激光投影仪",
        "英文品名": "Home Laser Projector",
        "品牌": "BrandA",
        "型号": "LP-1000",
        "HS编码": "8528690000",
        "用途": "家庭影音娱乐",
        "功能": "视频投射",
        "工作原理": "DLP数字光处理，激光光源",
        "材质": "塑料+金属+光学玻璃",
        "技术参数": "4K, 2000流明",
        "是否整机": "是",
        "是否含无线": "是",
        "是否含电池": "否",
        "数量": "10",
        "单位": "台",
        "单价": "500.00",
        "总价": "5000.00",
        "币种": "USD",
        "原产国": "日本",
        "毛重": "150.0",
        "净重": "120.0",
    },
    {
        "中文品名": "蓝牙音箱",
        "英文品名": "Bluetooth Speaker",
        "品牌": "BrandB",
        "型号": "BS-200",
        "HS编码": "8518220000",
        "用途": "便携音乐播放",
        "功能": "音频播放",
        "工作原理": "电动式扬声器",
        "材质": "塑料+金属网罩",
        "技术参数": "20W, 蓝牙5.2",
        "是否整机": "是",
        "是否含无线": "是",
        "是否含电池": "是",
        "数量": "20",
        "单位": "台",
        "单价": "50.00",
        "总价": "1000.00",
        "币种": "USD",
        "原产国": "中国",
        "毛重": "80.0",
        "净重": "70.0",
    },
]


# ──────────────────────────────────────────────
# CLI
# ──────────────────────────────────────────────


def print_risk_results(risks: list[dict]) -> None:
    """打印风险问题清单"""
    print("\n" + "=" * 60)
    print("风险问题清单")
    print("=" * 60)
    for r in risks:
        level = r.get("风险等级", r.get("level", "unknown"))
        title = r.get("问题说明", r.get("title", ""))
        position = r.get("问题位置", "")
        if position:
            print(f"  [{level}] {position}: {title}")
        else:
            print(f"  [{level}] {title}")
    print(f"\n共 {len(risks)} 个风险项")


def main() -> None:
    import argparse

    parser = argparse.ArgumentParser(description="报关资料表格生成脚本（增强版）")
    parser.add_argument("input_file", nargs="?", help="输入文件路径 (JSON/CSV/XLSX)")
    parser.add_argument("output_dir", nargs="?", help="输出目录")
    parser.add_argument(
        "output_format", nargs="?", choices=["xlsx", "csv", "json"], help="输出格式"
    )
    parser.add_argument(
        "--format",
        "-f",
        choices=["xlsx", "csv", "json"],
        default=None,
        help="输出格式（替代位置参数）",
    )
    args = parser.parse_args()

    # 兼容旧版 CLI: python script.py data.json output_dir xlsx
    if args.input_file:
        filepath = args.input_file
        if not os.path.exists(filepath):
            print(f"错误: 文件 '{filepath}' 不存在")
            sys.exit(1)
        data = load_data(filepath)
        if not data:
            print(f"错误: 无法从 '{filepath}' 加载数据")
            sys.exit(1)
    else:
        print("使用示例数据生成表格...\n")
        data = SAMPLE_DATA

    # 输出目录 & 格式
    output_dir = args.output_dir if args.output_dir else "./output"
    output_format = args.format or args.output_format or "xlsx"
    os.makedirs(output_dir, exist_ok=True)

    # ── 生成各类表格 ──
    summary = generate_summary_table(data)
    invoice, invoice_meta = generate_invoice(data)
    packing = generate_packing_list(data)
    elements, decl_source = generate_declaration_elements(data)
    questions = generate_pending_questions(data)
    risks, risk_source = detect_risks(data)

    # 输出引擎信息
    print(f"\n[信息] 申报要素来源: {decl_source}")
    print(f"[信息] 风险检测引擎: {risk_source}")

    # ── 构建输出 ──
    tables: dict[str, Any] = {
        "商品资料汇总": summary,
        "商业发票": invoice,
        "装箱单": packing,
        "申报要素明细": elements,
    }

    if output_format == "xlsx":
        output_path = os.path.join(output_dir, "declaration_tables.xlsx")
        save_as_xlsx(tables, output_path)
    elif output_format == "csv":
        for name, table in tables.items():
            fname = f"{name}.csv"
            save_as_csv(table, os.path.join(output_dir, fname))
    elif output_format == "json":
        output_path = os.path.join(output_dir, "declaration_tables.json")
        save_as_json(
            {
                "生成时间": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "数据来源": {
                    "申报要素": decl_source,
                    "风险检测": risk_source,
                },
                "商品资料汇总": summary,
                "商业发票": {"数据": invoice, "元数据": invoice_meta},
                "装箱单": packing,
                "申报要素明细": elements,
                "待确认问题": questions,
                "风险问题": risks,
            },
            output_path,
        )

    # ── 打印待确认问题 ──
    print("\n" + "=" * 60)
    print("待确认问题清单")
    print("=" * 60)
    for q in questions:
        print(f"  - {q}")
    print(f"\n共 {len(questions)} 个待确认问题")

    # ── 打印风险问题 ──
    print_risk_results(risks)

    # ── 保存问题和风险 ──
    problems: dict[str, Any] = {
        "生成时间": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "数据来源": {
            "申报要素": decl_source,
            "风险检测": risk_source,
        },
        "待确认问题": questions,
        "风险问题": risks,
    }
    save_as_json(problems, os.path.join(output_dir, "issues.json"))


if __name__ == "__main__":
    main()
